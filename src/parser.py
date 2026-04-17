"""
Excel parser for DQCP Master workbooks.

Reads the DQCP_Master sheet and optional lookup sheets (Data_Level,
Data_Sub_Level, Lookup Values) from each uploaded workbook.
"""
import hashlib
import json
from pathlib import Path

import openpyxl


# ── Lookup loaders ─────────────────────────────────────────────────────────────

def load_data_level_lookup(wb: openpyxl.Workbook) -> dict[int, dict]:
    """Return {Data_Level(int): {name, report_name, short_name, description}}."""
    if "Data_Level" not in wb.sheetnames:
        return {}
    rows = list(wb["Data_Level"].iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(v).strip() if v else "" for v in rows[0]]
    result = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            lvl = int(row[0])
        except (TypeError, ValueError):
            continue
        result[lvl] = {
            "data_level_name": _sv(row, header, "Data_Level_Name"),
            "data_level_report_name": _sv(row, header, "Data_Level_Report_Name"),
            "data_level_short_name": _sv(row, header, "Data_Level_Short_Name"),
        }
    return result


def load_data_sub_level_lookup(wb: openpyxl.Workbook) -> dict[int, dict]:
    """Return {Data_Sub_Level(int): {name, report_name}}."""
    if "Data_Sub_Level" not in wb.sheetnames:
        return {}
    rows = list(wb["Data_Sub_Level"].iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(v).strip() if v else "" for v in rows[0]]
    result = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            lvl = int(row[0])
        except (TypeError, ValueError):
            continue
        result[lvl] = {
            "data_sub_level_name": _sv(row, header, "Data_Sub_Level_Name"),
            "data_sub_level_report_name": _sv(row, header, "Data_Sub_Level_Report_Name"),
        }
    return result


def load_lookup_values(wb: openpyxl.Workbook) -> dict[str, list[dict]]:
    """Return {category: [{value, description}]} from the Lookup Values sheet."""
    if "Lookup Values" not in wb.sheetnames:
        return {}
    rows = list(wb["Lookup Values"].iter_rows(values_only=True))
    result = {}
    current_category = None
    for row in rows:
        first = str(row[0]).strip() if row[0] else ""
        second = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not first:
            continue
        # Category header rows have no numeric ID — they look like "DQCP_Status"
        if second in ("Description", "") and first not in ("Description",):
            # Detect a new category header
            if len(first) > 0 and not any(c.islower() for c in first[:4]):
                current_category = first
                result.setdefault(current_category, [])
                continue
        if current_category and first and first != "Description":
            result[current_category].append({"value": first, "description": second})
    return result


# ── Internal helpers ────────────────────────────────────────────────────────────

def _sv(row: tuple, header: list[str], col_name: str) -> str | None:
    """Safe value: get cell value by column name from a row tuple."""
    try:
        idx = header.index(col_name)
    except ValueError:
        return None
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    return str(v).strip() or None


def _sv_raw(row: tuple, header: list[str], col_name: str):
    """Like _sv but returns the raw value (for dates, numbers)."""
    try:
        idx = header.index(col_name)
    except ValueError:
        return None
    if idx >= len(row):
        return None
    return row[idx]


# ── Main parser ─────────────────────────────────────────────────────────────────

def parse_excel_files(file_paths: list[str], config: dict) -> list[dict]:
    """
    Parse a list of DQCP Master Excel file paths.

    Returns a list of dicts, one per DQCP checkpoint row from the
    DQCP_Master sheet, enriched with lookup descriptions.
    """
    xl_cfg = config["excel"]
    master_sheet = xl_cfg.get("master_sheet", "DQCP_Master")
    sync_statuses = {s.lower() for s in xl_cfg.get("sync_statuses", ["Active", "WIP"])}

    # Column name references from config
    id_col = xl_cfg["id_column"]
    title_col = xl_cfg["title_column"]
    description_col = xl_cfg["description_column"]
    pseudo_code_col = xl_cfg["pseudo_code_column"]
    status_col = xl_cfg["status_column"]
    is_approved_col = xl_cfg["is_approved_column"]
    rollout_col = xl_cfg["rollout_column"]
    data_level_col = xl_cfg["data_level_column"]
    data_sub_level_col = xl_cfg["data_sub_level_column"]
    sequence_col = xl_cfg["sequence_column"]
    data_element_col = xl_cfg["data_element_column"]
    table_name_col = xl_cfg["table_name_column"]
    column_name_col = xl_cfg["column_name_column"]
    sys_table_col = xl_cfg["sys_table_column"]
    sys_column_col = xl_cfg["sys_column_column"]
    start_date_col = xl_cfg["start_date_column"]
    end_date_col = xl_cfg["end_date_column"]
    last_modified_col = xl_cfg["last_modified_column"]
    resolved_date_col = xl_cfg["resolved_date_column"]
    comments_col = xl_cfg["comments_column"]
    questions_col = xl_cfg["questions_column"]
    question_status_col = xl_cfg["question_status_column"]
    change_history_col = xl_cfg["change_history_column"]
    resolution_col = xl_cfg["resolution_column"]

    results = []

    for file_path in file_paths:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Load lookups from this workbook
        dl_lookup = load_data_level_lookup(wb)
        dsl_lookup = load_data_sub_level_lookup(wb)

        if master_sheet not in wb.sheetnames:
            wb.close()
            continue

        ws = wb[master_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue

        # First row is the header
        header = [str(v).strip() if v is not None else "" for v in rows[0]]

        for row_number, row in enumerate(rows[1:], start=2):
            dqcp_id = _sv(row, header, id_col)
            if not dqcp_id:
                continue  # skip blank rows

            status = _sv(row, header, status_col) or ""
            if sync_statuses and status.lower() not in sync_statuses:
                # Still include in results but mark as excluded from push
                excluded = True
            else:
                excluded = False

            # Numeric lookups
            try:
                data_level = int(_sv_raw(row, header, data_level_col) or 0)
            except (TypeError, ValueError):
                data_level = None
            try:
                data_sub_level = int(_sv_raw(row, header, data_sub_level_col) or 0)
            except (TypeError, ValueError):
                data_sub_level = None

            # Enrich with lookup names
            dl_info = dl_lookup.get(data_level, {})
            dsl_info = dsl_lookup.get(data_sub_level, {})

            # Dates — store as ISO strings or None
            def _date(col_name):
                v = _sv_raw(row, header, col_name)
                if v is None:
                    return None
                try:
                    return v.date().isoformat() if hasattr(v, "date") else str(v)
                except Exception:
                    return str(v)

            row_dict = {
                # ── Identity ──────────────────────────────────────────────
                "checkpoint_key": f"{Path(file_path).name}::{dqcp_id}",
                "file_path": str(file_path),
                "sheet_name": master_sheet,
                "row_number": row_number,
                "dqcp_id": dqcp_id,
                "data_level": data_level,
                "data_sub_level": data_sub_level,
                "sequence_number": _sv(row, header, sequence_col),
                "data_element": _sv(row, header, data_element_col),

                # ── Lookup enrichment ─────────────────────────────────────
                "data_level_name": dl_info.get("data_level_name"),
                "data_level_report_name": dl_info.get("data_level_report_name"),
                "data_sub_level_name": dsl_info.get("data_sub_level_name"),
                "data_sub_level_report_name": dsl_info.get("data_sub_level_report_name"),

                # ── Content ───────────────────────────────────────────────
                "checkpoint_name": _sv(row, header, title_col),    # ADO title
                "dqcp_title": _sv(row, header, title_col),
                "dqcp_description": _sv(row, header, description_col),
                "dqcp_pseudo_code": _sv(row, header, pseudo_code_col),

                # ── Source mapping ────────────────────────────────────────
                "table_name": _sv(row, header, table_name_col),
                "column_name": _sv(row, header, column_name_col),
                "sys_table_name": _sv(row, header, sys_table_col),
                "sys_column_name": _sv(row, header, sys_column_col),

                # ── Status flags ──────────────────────────────────────────
                "status": status,
                "is_approved": _sv(row, header, is_approved_col),
                "rollout": _sv(row, header, rollout_col),

                # ── Dates ─────────────────────────────────────────────────
                "start_date": _date(start_date_col),
                "end_date": _date(end_date_col),
                "last_modified_date": _date(last_modified_col),
                "resolved_date": _date(resolved_date_col),

                # ── Open items / audit ────────────────────────────────────
                "dqcp_comments": _sv(row, header, comments_col),
                "dqcp_questions": _sv(row, header, questions_col),
                "question_status": _sv(row, header, question_status_col),
                "change_history": _sv(row, header, change_history_col),
                "resolution": _sv(row, header, resolution_col),

                # ── Sync control ──────────────────────────────────────────
                "excluded_from_sync": excluded,
            }
            row_dict["field_hash"] = generate_field_hash(row_dict)
            results.append(row_dict)

        wb.close()

    return results


def generate_field_hash(row_dict: dict) -> str:
    """
    SHA-256 hash over the fields that matter for change detection.

    Fields that drive a re-push when changed:
      status, is_approved, rollout, dqcp_description, dqcp_pseudo_code,
      dqcp_comments, resolution, end_date
    """
    key_fields = {
        "status": row_dict.get("status"),
        "is_approved": row_dict.get("is_approved"),
        "rollout": row_dict.get("rollout"),
        "dqcp_description": row_dict.get("dqcp_description"),
        "dqcp_pseudo_code": row_dict.get("dqcp_pseudo_code"),
        "dqcp_comments": row_dict.get("dqcp_comments"),
        "resolution": row_dict.get("resolution"),
        "end_date": row_dict.get("end_date"),
    }
    serialised = json.dumps(key_fields, sort_keys=True, default=str)
    return hashlib.sha256(serialised.encode("utf-8")).hexdigest()
